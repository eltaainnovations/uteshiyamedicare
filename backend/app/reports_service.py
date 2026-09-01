"""Builds the 4 report .xlsx workbooks. No ERPNext calls live here — every
data fetch reuses orders_service.py's existing functions. This module's
only job is orchestration + spreadsheet formatting.
"""

from datetime import date
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from . import orders_service
from .schemas import ReportType

_TOP_PRODUCTS_LIMIT = 10


def _write_sheet(ws: Worksheet, headers: list[str], rows: list[list[Any]], *, widths: list[int] | None = None) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append(row)
    for i, width in enumerate(widths or [max(12, len(h) + 2) for h in headers], start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width


async def _build_sales_sheet(ws: Worksheet, from_date: date, to_date: date) -> None:
    orders = await orders_service.list_orders_in_range(str(from_date), str(to_date))
    rows = [
        [o["transaction_date"], o["name"], o["customer_name"] or o["customer"], o["grand_total"]] for o in orders
    ]
    total = sum(o["grand_total"] for o in orders)
    rows.append(["", "", "Total", total])
    _write_sheet(ws, ["Date", "Order ID", "Distributor", "Amount"], rows, widths=[14, 24, 30, 16])


async def _build_orders_sheet(ws: Worksheet, from_date: date, to_date: date) -> None:
    orders = await orders_service.list_orders_in_range(str(from_date), str(to_date))
    rows = [
        [
            o["name"],
            o["customer_name"] or o["customer"],
            o["item_count"],
            o["grand_total"],
            o["status"],
            o["delivery_date"],
        ]
        for o in orders
    ]
    _write_sheet(
        ws,
        ["Order ID", "Distributor", "Items", "Value", "Status", "Delivery Date"],
        rows,
        widths=[24, 30, 10, 16, 20, 16],
    )


async def _build_top_products_sheet(ws: Worksheet, from_date: date, to_date: date) -> None:
    rollup = await orders_service.item_rollup_in_range(str(from_date), str(to_date))
    top = rollup[:_TOP_PRODUCTS_LIMIT]
    rows = [[r["item_name"] or r["item_code"], r["item_code"], r["qty"], r["revenue"]] for r in top]
    _write_sheet(ws, ["Product Name", "Item Code", "Total Qty", "Total Revenue"], rows, widths=[36, 20, 14, 18])


async def _build_distributor_performance_sheet(ws: Worksheet, from_date: date, to_date: date) -> None:
    ranked = await orders_service.distributor_performance_in_range(str(from_date), str(to_date))
    rows = [[r["distributor"], r["order_count"], r["total_value"]] for r in ranked]
    _write_sheet(ws, ["Distributor", "Order Count", "Total Value"], rows, widths=[32, 14, 18])


_SHEET_BUILDERS = {
    "sales": ("Sales & Revenue", _build_sales_sheet),
    "orders": ("Orders", _build_orders_sheet),
    "top_products": ("Top Products", _build_top_products_sheet),
    "distributor_performance": ("Distributor Performance", _build_distributor_performance_sheet),
}


async def generate_report(report_type: ReportType, from_date: date, to_date: date) -> bytes:
    sheet_title, builder = _SHEET_BUILDERS[report_type]
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    await builder(ws, from_date, to_date)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
